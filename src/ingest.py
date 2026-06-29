"""
Ingestion adapters — one per feed STANDARD, not per supplier. Because most
promotional suppliers deliver via BMEcat or Promidata, two adapters already
cover the long tail; CSV/Excel adapters handle the rest via per-supplier config.

Every adapter returns records in the same intermediate shape:
    supplier_sku, ean, name, manufacturer, category(raw), material, colors[],
    price_tiers[{qty, amount, currency}], min_order_qty, print_methods[],
    print_area, weight, weight_unit, eco(bool), image_url, stock_qty,
    lead_time_days, description
"""
import csv, json, os
import xml.etree.ElementTree as ET


def _num(v):
    if v in (None, ""):
        return None
    try:
        return float(str(v).replace(",", "."))
    except ValueError:
        return None


def _split(v, sep):
    if not v:
        return []
    return [x.strip() for x in str(v).split(sep) if x.strip()]


# ------------------------------------------------------------------- BMEcat 1.2
def read_bmecat(path, cfg):
    tree = ET.parse(path)
    out = []
    for art in tree.iter("ARTICLE"):
        det = art.find("ARTICLE_DETAILS")
        feats = {f.findtext("FNAME"): f.findtext("FVALUE") for f in art.iter("FEATURE")}
        tiers = []
        for pr in art.iter("ARTICLE_PRICE"):
            tiers.append({"qty": int(_num(pr.findtext("LOWER_BOUND")) or 0),
                          "amount": _num(pr.findtext("PRICE_AMOUNT")),
                          "currency": pr.findtext("PRICE_CURRENCY") or cfg.get("currency_default", "EUR")})
        out.append({
            "supplier_sku": art.findtext("SUPPLIER_AID"),
            "ean": (det.findtext("EAN") or "").strip() if det is not None else "",
            "name": det.findtext("DESCRIPTION_SHORT") if det is not None else None,
            "manufacturer": det.findtext("MANUFACTURER_NAME") if det is not None else None,
            "category": feats.get("Warengruppe"),
            "material": feats.get("Material"),
            "colors": _split(feats.get("Farben"), ","),
            "price_tiers": tiers,
            "min_order_qty": _num(feats.get("Mindestmenge")),
            "print_methods": _split(feats.get("Veredelung"), ","),
            "print_area": feats.get("Werbeflaeche"),
            "weight": _num(feats.get("GewichtG")), "weight_unit": "g",
            "eco": (feats.get("Nachhaltig", "").lower() == "ja"),
            "image_url": (art.findtext("MIME_INFO/MIME/MIME_SOURCE") or "").strip(),
            "stock_qty": _num(art.findtext("STOCK")),
            "lead_time_days": _num(det.findtext("DELIVERY_TIME") if det is not None else None),
            "description": det.findtext("DESCRIPTION_LONG") if det is not None else None,
        })
    return out


# -------------------------------------------------------------------- Promidata
def read_promidata(path, cfg):
    with open(path, encoding="utf-8") as f:
        payload = json.load(f)
    out = []
    for p in payload.get("Products", []):
        d = p.get("ProductDetails", {})
        imgs = p.get("ImageList", [])
        out.append({
            "supplier_sku": d.get("SupplierAID"),
            "ean": (d.get("EAN") or "").strip(),
            "name": (d.get("Name") or {}).get("de"),
            "manufacturer": d.get("Manufacturer"),
            "category": d.get("Category"),
            "material": (d.get("Material") or {}).get("de"),
            "colors": [(c.get("Color") or {}).get("de") for c in p.get("ChildProducts", []) if c.get("Color")],
            "price_tiers": [{"qty": int(s.get("Scale") or 0), "amount": _num(s.get("Price")),
                             "currency": s.get("Currency", cfg.get("currency_default", "EUR"))}
                            for s in p.get("PriceList", [])],
            "min_order_qty": _num(d.get("MinimumOrderQuantity")),
            "print_methods": d.get("PrintingTechniques") or [],
            "print_area": d.get("PrintingDimensions"),
            "weight": _num(d.get("Weight")), "weight_unit": d.get("WeightUnit", "g"),
            "eco": bool(d.get("Sustainable")),
            "image_url": (imgs[0]["Url"] if imgs else ""),
            "stock_qty": _num((p.get("Stock") or {}).get("Quantity")),
            "lead_time_days": _num(d.get("DeliveryTime")),
            "description": (d.get("Description") or {}).get("de"),
        })
    return out


# -------------------------------------------------------------- CSV / Excel rows
def _rows_csv(path, cfg):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f, delimiter=cfg.get("options", {}).get("delimiter", ",")))


def _rows_excel(path, cfg):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = cfg.get("options", {}).get("sheet")
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    return [{str(h): v for h, v in zip(header, r)} for r in rows[1:]]


def _from_tabular(rows, cfg):
    fm = cfg["field_map"]
    out = []
    for row in rows:
        rec = {}
        for target, col in fm.items():
            if target in ("weight_unit",):
                continue
            rec[target] = row.get(col)
        rec["colors"] = _split(rec.get("colors"), cfg.get("colors_sep", ","))
        rec["print_methods"] = _split(rec.get("print_methods"), cfg.get("print_sep", ","))
        rec["min_order_qty"] = _num(rec.get("min_order_qty"))
        rec["stock_qty"] = _num(rec.get("stock_qty"))
        rec["lead_time_days"] = _num(rec.get("lead_time_days"))
        rec["weight"] = _num(rec.get("weight_kg"))
        rec["weight_unit"] = fm.get("weight_unit", "kg")
        rec.pop("weight_kg", None)
        rec["eco"] = False
        tiers = []
        for t in cfg.get("price_tiers_csv", []):
            amt = _num(row.get(t["field"]))
            tiers.append({"qty": t["qty"], "amount": amt, "currency": cfg.get("currency_default", "EUR")})
        rec["price_tiers"] = tiers
        out.append(rec)
    return out


def read_csv(path, cfg):
    return _from_tabular(_rows_csv(path, cfg), cfg)


def read_excel(path, cfg):
    return _from_tabular(_rows_excel(path, cfg), cfg)


ADAPTERS = {"bmecat": read_bmecat, "promidata": read_promidata, "csv": read_csv, "excel": read_excel}


def ingest(cfg, root):
    fmt = cfg["format"]
    if fmt not in ADAPTERS:
        raise ValueError(f"No adapter for format '{fmt}'")
    return ADAPTERS[fmt](os.path.join(root, cfg["source"]), cfg)
